"""
Job Scraper — Ashley Ives
Profile: Chemistry PhD / Mass Spectrometry / Proteomics / Protein Biochemistry / R

Sources (all free, API keys required — see README.md):
  1. Adzuna API      — broad US + UK/CA/AU/NZ jobs (250 req/day free)
  2. USAJobs API     — federal government positions
  3. Jooble API      — global job aggregator
  4. jobs.ac.uk      — UK academic / research positions (HTML + JSON-LD)
  5. Curated links   — one-click search URLs for manually checking boards

Location tiers:
  Onsite/Hybrid  → US only;   WA = A, other US = B
  Remote         → US + english-speaking countries; all = B (WA = A)
"""

import os, re, json, time, logging, hashlib
from datetime import datetime
from pathlib import Path
from urllib.parse import urlencode, quote_plus

import certifi, feedparser, requests
import pandas as pd
from bs4 import BeautifulSoup

import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# ── Paths ──────────────────────────────────────────────────────────────────
BASE_DIR   = Path(__file__).parent
OUTPUT_DIR = BASE_DIR / "output"
OUTPUT_DIR.mkdir(exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(BASE_DIR / "scraper.log", encoding="utf-8"),
        logging.StreamHandler(),
    ],
)
log = logging.getLogger(__name__)

# ── API credentials (set env vars or paste values here) ───────────────────
# Adzuna  — register free at https://developer.adzuna.com/
ADZUNA_APP_ID  = os.environ.get("ADZUNA_APP_ID",  "")
ADZUNA_APP_KEY = os.environ.get("ADZUNA_APP_KEY", "")

# USAJobs — register free at https://developer.usajobs.gov/
USAJOBS_KEY   = os.environ.get("USAJOBS_API_KEY", "")
USAJOBS_EMAIL = os.environ.get("USAJOBS_EMAIL",   "")

# Jooble  — register free at https://jooble.org/api/about
JOOBLE_KEY = os.environ.get("JOOBLE_API_KEY", "")

# SerpAPI — Google Jobs (covers Indeed, BioSpace, LinkedIn, Glassdoor)
SERPAPI_KEY = os.environ.get("SERPAPI_KEY", "1210831b18136831290b9082654a2c8b402d00de7bed4903497e8ff2c23ff651")

# ── Search queries (tailored to resume) ───────────────────────────────────
QUERIES = [
    "mass spectrometry",
    "proteomics",
    "top-down proteomics",
    "analytical chemistry research scientist",
    "protein biochemistry scientist",
    "mass spectrometry scientist",
    "research scientist biochemistry",
]

# ── Geography helpers ──────────────────────────────────────────────────────
WA_SIGNALS = {
    "washington state", "seattle", "tacoma", "bellevue", "spokane",
    "redmond", "kirkland", "bothell", "renton", "everett", "olympia",
    "bellingham", "yakima", "kennewick", "pasco", "richland",
    "federal way", "shoreline", "burien", "lacey", "marysville",
}
US_ABBREVS = {
    "AL","AK","AZ","AR","CA","CO","CT","DE","FL","GA","HI","ID","IL","IN",
    "IA","KS","KY","LA","ME","MD","MA","MI","MN","MS","MO","MT","NE","NV",
    "NH","NJ","NM","NY","NC","ND","OH","OK","OR","PA","RI","SC","SD","TN",
    "TX","UT","VT","VA","WA","WV","WI","WY","DC",
}
ENGLISH_COUNTRIES = {
    "united states","usa","us","canada","united kingdom","uk","great britain",
    "england","scotland","wales","ireland","australia","new zealand",
}
REMOTE_SIGNALS = {"remote","fully remote","work from home","wfh","telework","telecommute","distributed"}
HYBRID_SIGNALS = {"hybrid","flex schedule","flexible location"}

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "en-US,en;q=0.9",
}

# ── Utilities ──────────────────────────────────────────────────────────────

def safe_get(url: str, verify: bool = True, **kw) -> requests.Response | None:
    for attempt in range(3):
        try:
            time.sleep(1.2)
            r = requests.get(url, headers=HEADERS, timeout=20,
                             verify=certifi.where() if verify else False, **kw)
            if r.status_code == 200:
                return r
            log.debug("HTTP %s — %s", r.status_code, url)
        except Exception as exc:
            log.debug("Attempt %d failed for %s: %s", attempt + 1, url, exc)
    log.warning("All retries failed: %s", url)
    return None


def detect_work_type(text: str) -> str:
    low = text.lower()
    if any(s in low for s in REMOTE_SIGNALS):
        return "Remote"
    if any(s in low for s in HYBRID_SIGNALS):
        return "Hybrid"
    return "Onsite"


def _is_wa(loc: str) -> bool:
    ll = loc.lower()
    return any(s in ll for s in WA_SIGNALS) or re.search(r'\bwa\b', ll) is not None


def _is_us(loc: str) -> bool:
    tokens = re.split(r"[\s,]+", loc.upper())
    return any(t in US_ABBREVS for t in tokens) or "USA" in loc.upper() or "UNITED STATES" in loc.upper()


def _is_english(loc: str) -> bool:
    ll = loc.lower()
    return any(c in ll for c in ENGLISH_COUNTRIES) or _is_us(loc)


def location_tier(loc: str, work_type: str) -> str:
    if work_type == "Remote":
        if not loc.strip() or loc.lower() in ("remote", "anywhere", "worldwide"):
            return "B"   # unspecified remote → allow but tier B
        if _is_wa(loc):
            return "A"
        if _is_english(loc):
            return "B"
        return "SKIP"
    # Onsite / Hybrid — US only
    if _is_wa(loc):
        return "A"
    if _is_us(loc):
        return "B"
    return "SKIP"


def extract_salary(text: str) -> str:
    pat = (
        r"\$[\d,]+(?:\.\d+)?(?:\s*[–\-]\s*\$[\d,]+(?:\.\d+)?)?(?:\s*/\s*(?:yr|year|hr|hour|annually|k))?"
        r"|£[\d,]+(?:\.\d+)?(?:\s*[–\-]\s*£[\d,]+(?:\.\d+)?)?"
        r"|[\d,]+\s*[–\-]\s*[\d,]+\s*(?:USD|CAD|GBP|AUD)"
    )
    m = re.search(pat, text, re.IGNORECASE)
    return m.group(0).strip() if m else "Not listed"


def dedupe(jobs: list[dict]) -> list[dict]:
    seen, out = set(), []
    for j in jobs:
        key = hashlib.md5(
            f"{j.get('Job Title','').lower().strip()}"
            f"{j.get('Company','').lower().strip()}"
            f"{j.get('Location','').lower().strip()}".encode()
        ).hexdigest()
        if key not in seen:
            seen.add(key)
            out.append(j)
    return out


def _row(*, source, title, company, salary, location, work_type, link,
         docs="Resume, Cover Letter") -> dict | None:
    tier = location_tier(location, work_type)
    if tier == "SKIP":
        return None
    return {
        "Source":             source,
        "Job Title":          title,
        "Company":            company,
        "Salary":             salary,
        "Location":           location,
        "Work Type":          work_type,
        "Location Tier":      tier,
        "Application Link":   link,
        "Documents Required": docs,
        "Date Scraped":       datetime.now().strftime("%Y-%m-%d"),
    }


# ── Source 1: Adzuna API ──────────────────────────────────────────────────
# Free: 250 requests/day — register at https://developer.adzuna.com/

ADZUNA_COUNTRIES = [
    ("us", "Washington"),   # US — filter WA
    ("us", ""),             # US — all remote
    ("gb", ""),             # UK
    ("ca", ""),             # Canada
    ("au", ""),             # Australia
    ("nz", ""),             # New Zealand
]


def scrape_adzuna(keyword: str) -> list[dict]:
    if not ADZUNA_APP_ID or not ADZUNA_APP_KEY:
        return []
    jobs = []
    for country, where in ADZUNA_COUNTRIES:
        params = {
            "app_id": ADZUNA_APP_ID,
            "app_key": ADZUNA_APP_KEY,
            "results_per_page": 50,
            "what": keyword,
            "content-type": "application/json",
        }
        if where:
            params["where"] = where
        url = f"https://api.adzuna.com/v1/api/jobs/{country}/search/1?{urlencode(params)}"
        r = safe_get(url)
        if not r:
            continue
        try:
            data = r.json()
        except Exception:
            continue
        for item in data.get("results", []):
            loc      = item.get("location", {}).get("display_name", "")
            wt       = detect_work_type(item.get("title","") + " " + item.get("description",""))
            sal_min  = item.get("salary_min")
            sal_max  = item.get("salary_max")
            if sal_min and sal_max:
                sal = f"${sal_min:,.0f}–${sal_max:,.0f}"
            elif sal_min:
                sal = f"${sal_min:,.0f}+"
            else:
                sal = extract_salary(item.get("description",""))
            row = _row(
                source=f"Adzuna ({country.upper()})",
                title=item.get("title",""),
                company=item.get("company",{}).get("display_name","See listing"),
                salary=sal,
                location=loc,
                work_type=wt,
                link=item.get("redirect_url",""),
            )
            if row:
                jobs.append(row)
        time.sleep(0.5)
    log.info("Adzuna [%s]: %d jobs", keyword, len(jobs))
    return jobs


# ── Source 2: USAJobs API ─────────────────────────────────────────────────
# Free — register at https://developer.usajobs.gov/

def scrape_usajobs(keyword: str) -> list[dict]:
    if not USAJOBS_KEY or not USAJOBS_EMAIL:
        return []
    jobs = []
    url = ("https://data.usajobs.gov/api/search?"
           + urlencode({"Keyword": keyword, "ResultsPerPage": 50, "Fields": "Min"}))
    try:
        r = requests.get(url, headers={
            **HEADERS,
            "Host": "data.usajobs.gov",
            "User-Agent": USAJOBS_EMAIL,
            "Authorization-Key": USAJOBS_KEY,
        }, timeout=20, verify=certifi.where())
        items = r.json().get("SearchResult", {}).get("SearchResultItems", [])
    except Exception as exc:
        log.error("USAJobs error: %s", exc)
        return []

    for item in items:
        pos  = item.get("MatchedObjectDescriptor", {})
        locs = pos.get("PositionLocation", [{}])
        loc  = "; ".join(f"{l.get('CityName','')}, {l.get('CountrySubDivisionCode','')}" for l in locs)
        rem  = pos.get("PositionRemuneration", [{}])
        lo   = rem[0].get("MinimumRange","") if rem else ""
        hi   = rem[0].get("MaximumRange","") if rem else ""
        sal  = f"${lo}–${hi}/yr" if lo and hi else "Not listed"
        wt   = detect_work_type(pos.get("PositionTitle","") + " " + pos.get("QualificationSummary",""))
        row  = _row(
            source="USAJobs.gov",
            title=pos.get("PositionTitle",""),
            company=pos.get("OrganizationName",""),
            salary=sal, location=loc, work_type=wt,
            link=pos.get("PositionURI",""),
            docs="Resume, Transcripts (federal standard)",
        )
        if row:
            jobs.append(row)
    log.info("USAJobs [%s]: %d jobs", keyword, len(jobs))
    return jobs


# ── Source 3: Jooble API ──────────────────────────────────────────────────
# Free — register at https://jooble.org/api/about

JOOBLE_SEARCHES = [
    ("Seattle, WA", True),
    ("Washington State", True),
    ("Remote", False),
    ("London, UK", False),
    ("Toronto, Canada", False),
]


def scrape_jooble(keyword: str) -> list[dict]:
    if not JOOBLE_KEY:
        return []
    jobs = []
    url = f"https://jooble.org/api/{JOOBLE_KEY}"
    for location, _ in JOOBLE_SEARCHES:
        payload = json.dumps({"keywords": keyword, "location": location, "page": 1})
        try:
            time.sleep(1)
            r = requests.post(url, data=payload,
                              headers={**HEADERS, "Content-Type": "application/json"},
                              timeout=20, verify=certifi.where())
            data = r.json()
        except Exception as exc:
            log.warning("Jooble error [%s / %s]: %s", keyword, location, exc)
            continue
        for item in data.get("jobs", []):
            loc = item.get("location", location)
            wt  = detect_work_type(item.get("title","") + " " + item.get("snippet","") + " " + loc)
            row = _row(
                source="Jooble",
                title=item.get("title",""),
                company=item.get("company","See listing"),
                salary=item.get("salary","") or extract_salary(item.get("snippet","")),
                location=loc,
                work_type=wt,
                link=item.get("link",""),
            )
            if row:
                jobs.append(row)
    log.info("Jooble [%s]: %d jobs", keyword, len(jobs))
    return jobs


# ── Source 4: jobs.ac.uk (UK academic / research) ─────────────────────────
# HTML scrape + JSON-LD extraction — no key needed

def scrape_jobsacuk(keyword: str) -> list[dict]:
    jobs = []
    url = "https://www.jobs.ac.uk/search/?" + urlencode({"keywords": keyword})
    r = safe_get(url)
    if not r:
        return []
    soup = BeautifulSoup(r.text, "lxml")

    # Method A: JSON-LD structured data
    for script in soup.find_all("script", type="application/ld+json"):
        try:
            data = json.loads(script.string or "")
            items = data if isinstance(data, list) else [data]
            for item in items:
                if item.get("@type") == "JobPosting":
                    loc = (item.get("jobLocation") or {}).get("address", {})
                    if isinstance(loc, str):
                        loc_str = loc
                    else:
                        loc_str = ", ".join(filter(None, [
                            loc.get("addressLocality",""),
                            loc.get("addressRegion",""),
                            loc.get("addressCountry",""),
                        ]))
                    sal_data = item.get("baseSalary", {}).get("value", {})
                    sal = (
                        f"{sal_data.get('currency','')}{sal_data.get('minValue','')}–"
                        f"{sal_data.get('maxValue','')}"
                        if isinstance(sal_data, dict) else "Not listed"
                    )
                    wt = detect_work_type(item.get("title","") + " " + item.get("description",""))
                    row = _row(
                        source="jobs.ac.uk",
                        title=item.get("title",""),
                        company=item.get("hiringOrganization",{}).get("name","See listing"),
                        salary=sal, location=loc_str, work_type=wt,
                        link=item.get("url", url),
                        docs="CV, Cover Letter, References (academic standard)",
                    )
                    if row:
                        jobs.append(row)
        except Exception:
            pass

    # Method B: visible HTML cards (if JSON-LD empty)
    if not jobs:
        kw_lower = keyword.lower().split()
        for card in soup.select("div.j-search-result__item, article.j-search-result__item, li[class*='result']"):
            text = card.get_text(" ", strip=True)
            if not any(k in text.lower() for k in kw_lower):
                continue
            title_el   = card.select_one("a.j-search-result__anchor, h3 a, h2 a")
            company_el = card.select_one(".j-search-result__name, .employer")
            loc_el     = card.select_one(".j-search-result__location, .location")
            title   = title_el.get_text(strip=True) if title_el else text[:80]
            href    = title_el["href"] if title_el and title_el.get("href") else url
            if href.startswith("/"):
                href = "https://www.jobs.ac.uk" + href
            company = company_el.get_text(strip=True) if company_el else "See listing"
            loc     = loc_el.get_text(strip=True) if loc_el else "United Kingdom"
            wt      = detect_work_type(text)
            row     = _row(source="jobs.ac.uk", title=title, company=company,
                           salary=extract_salary(text), location=loc, work_type=wt,
                           link=href, docs="CV, Cover Letter, References")
            if row:
                jobs.append(row)

    log.info("jobs.ac.uk [%s]: %d jobs", keyword, len(jobs))
    return jobs


# ── Source 5: SerpAPI — Google Jobs ──────────────────────────────────────
# Aggregates from Indeed, BioSpace, LinkedIn, Glassdoor & more.
# Free tier: 100 searches/month — register at https://serpapi.com/

SERP_LOCATIONS = [
    ("Seattle, Washington, United States", ""),
    ("Washington State, United States",    ""),
    ("",                                   "work_from_home"),   # remote jobs
    ("United Kingdom",                     ""),
    ("Canada",                             ""),
]


def scrape_serpapi(keyword: str) -> list[dict]:
    if not SERPAPI_KEY:
        return []
    jobs = []
    for loc, chips in SERP_LOCATIONS:
        params = {
            "engine":  "google_jobs",
            "q":       keyword,
            "hl":      "en",
            "api_key": SERPAPI_KEY,
        }
        if loc:
            params["location"] = loc
        if chips:
            params["ltype"] = chips
        r = safe_get("https://serpapi.com/search?" + urlencode(params))
        if not r:
            continue
        try:
            data = r.json()
        except Exception:
            continue

        for item in data.get("jobs_results", []):
            title   = item.get("title", "")
            company = item.get("company_name", "See listing")
            location_raw = item.get("location", loc if loc else "Remote")

            # Salary — SerpAPI nests it inside detected_extensions
            ext = item.get("detected_extensions", {})
            sal = ext.get("salary", "") or extract_salary(item.get("description", ""))

            # Work type
            schedule = ext.get("schedule_type", "")
            wt = detect_work_type(title + " " + location_raw + " " + schedule + " " + item.get("description", ""))

            # Application link — prefer apply_options[0], fall back to share link
            apply_opts = item.get("apply_options", [])
            link = apply_opts[0].get("link", "") if apply_opts else item.get("share_link", "")

            # Documents — check if posting mentions specific requirements
            desc = item.get("description", "")
            docs = "Resume, Cover Letter"
            if any(w in desc.lower() for w in ["cover letter", "cv", "curriculum vitae"]):
                docs = "Resume/CV, Cover Letter"
            if any(w in desc.lower() for w in ["reference", "letter of recommendation"]):
                docs += ", References"

            # Source label — show which board it came from if available
            source_board = apply_opts[0].get("title", "Google Jobs") if apply_opts else "Google Jobs"
            source = f"Google Jobs ({source_board})"

            row = _row(
                source=source,
                title=title,
                company=company,
                salary=sal or "Not listed",
                location=location_raw,
                work_type=wt,
                link=link,
                docs=docs,
            )
            if row:
                jobs.append(row)
        time.sleep(1)

    log.info("SerpAPI/Google Jobs [%s]: %d jobs", keyword, len(jobs))
    return jobs


# ── Job board search links (separate sheet, NOT in All Jobs) ──────────────
# These are one-click search URLs for boards that require a browser.
# Shown on the "Job Board Links" sheet — not mixed into real job listings.

SEARCH_LINK_BOARDS = [
    # (Board name, URL template with {qe} placeholder, Location label, Notes)
    ("Indeed — Washington State",
     "https://www.indeed.com/jobs?q={qe}&l=Washington+State&sort=date&fromage=14",
     "Washington, USA", "Filter by date posted; check daily for new listings"),
    ("Indeed — Remote",
     "https://www.indeed.com/jobs?q={qe}&l=Remote&sort=date&fromage=14",
     "Remote", "Remote-only filter; worldwide English listings appear here"),
    ("LinkedIn — Washington State",
     "https://www.linkedin.com/jobs/search/?keywords={qe}&location=Washington%2C+United+States&f_TPR=r604800",
     "Washington, USA", "Past-week filter applied; sign in for full details"),
    ("LinkedIn — Remote",
     "https://www.linkedin.com/jobs/search/?keywords={qe}&f_WT=2&f_TPR=r604800",
     "Remote", "Remote filter + past week; worldwide English listings"),
    ("Nature Careers",
     "https://www.nature.com/naturecareers/jobs/search?keywords={qe}",
     "Various", "Academic & industry science roles globally"),
    ("Science Careers (AAAS)",
     "https://jobs.sciencecareers.org/jobs/?keywords={qe}",
     "Various", "Strong for academic/postdoc science roles"),
    ("BioSpace",
     "https://www.biospace.com/jobs?q={qe}",
     "Various", "Life sciences & biotech industry focus"),
    ("Glassdoor — Washington",
     "https://www.glassdoor.com/Job/washington-{qe_dashed}-jobs-SRCH_IL.0,10_IS2019.htm",
     "Washington, USA", "Includes salary estimates and reviews"),
    ("USAJobs.gov — Washington",
     "https://www.usajobs.gov/Search/Results?k={qe}&l=Washington&p=1",
     "Washington, USA", "All federal government positions; also try l=Remote"),
    ("ASMS Job Bank",
     "https://www.asms.org/careers/job-bank",
     "Various", "American Society for Mass Spectrometry — niche, highly relevant"),
    ("HigherEdJobs",
     "https://www.higheredjobs.com/search/advanced_action.cfm?Keyword={qe}",
     "Various", "University & research institution positions"),
    ("PNNL Careers",
     "https://careers.pnnl.gov/jobs?keywords={qe}&location=",
     "Richland / Seattle, WA", "Pacific Northwest National Laboratory — strong MS/proteomics group"),
    ("Fred Hutchinson Cancer Center",
     "https://www.fredhutch.org/en/jobs/find-a-job.html?term={qe}",
     "Seattle, WA", "Leading cancer research centre in Seattle"),
    ("Allen Institute",
     "https://alleninstitute.org/careers/?search={qe}",
     "Seattle, WA", "Brain & cell science research; Seattle headquarters"),
    ("UW Medicine / UW Careers",
     "https://careers.uw.edu/en-us/search/?search={qe}&location=Seattle",
     "Seattle, WA", "University of Washington research & staff positions"),
    ("jobs.ac.uk — UK",
     "https://www.jobs.ac.uk/search/?keywords={qe}",
     "United Kingdom", "All UK academic/research roles; good for remote-eligible"),
]

LINK_QUERIES = [
    "mass spectrometry",
    "proteomics",
    "analytical chemistry scientist",
]


def build_search_links() -> list[dict]:
    """Return rows for the Job Board Links sheet."""
    rows = []
    seen = set()
    for q in LINK_QUERIES:
        qe      = quote_plus(q)
        qe_dash = q.replace(" ", "-")
        for board, url_tmpl, loc, notes in SEARCH_LINK_BOARDS:
            url = url_tmpl.replace("{qe}", qe).replace("{qe_dashed}", qe_dash)
            key = (board, q)
            if key in seen:
                continue
            seen.add(key)
            rows.append({
                "Job Board":    board,
                "Search Query": q,
                "Location":     loc,
                "Link":         url,
                "Notes":        notes,
            })
    return rows


# ── Excel output ───────────────────────────────────────────────────────────

COLUMNS = [
    "Source", "Job Title", "Company", "Salary",
    "Location", "Work Type", "Location Tier",
    "Application Link", "Documents Required", "Date Scraped",
]


def _style_sheet(ws, link_col_letter: str, fill_fn, bold_header: bool = True):
    """Apply shared formatting to a worksheet."""
    from openpyxl.styles import Font, Alignment
    if bold_header:
        for cell in ws[1]:
            cell.font = Font(bold=True)
    for row in ws.iter_rows(min_row=2):
        fill = fill_fn(row)
        for cell in row:
            if fill:
                cell.fill = fill
            cell.alignment = Alignment(wrap_text=False, vertical="top")
    # Hyperlink the link column
    link_idx = ws[f"{link_col_letter}1"].column
    for row in ws.iter_rows(min_row=2, min_col=link_idx, max_col=link_idx):
        cell = row[0]
        if cell.value and str(cell.value).startswith("http"):
            cell.hyperlink = str(cell.value)
            cell.font = Font(color="0563C1", underline="single")


def save_excel(jobs: list[dict], search_links: list[dict]) -> Path:
    from openpyxl.styles import PatternFill, Font

    green  = PatternFill("solid", fgColor="C6EFCE")
    yellow = PatternFill("solid", fgColor="FFEB9C")
    blue   = PatternFill("solid", fgColor="DDEEFF")

    date_str = datetime.now().strftime("%Y-%m-%d")
    out_path = OUTPUT_DIR / f"jobs_{date_str}.xlsx"

    # ── Sheet 1: All Jobs ──────────────────────────────────────────────────
    df_jobs = pd.DataFrame(jobs, columns=COLUMNS) if jobs else pd.DataFrame(columns=COLUMNS)
    if not df_jobs.empty:
        df_jobs["_s"] = df_jobs["Location Tier"].map({"A": 0, "B": 1}).fillna(2)
        df_jobs = df_jobs.sort_values(["_s", "Source", "Job Title"]).drop(columns=["_s"])

    # ── Sheet 2: Job Board Links ───────────────────────────────────────────
    link_cols = ["Job Board", "Search Query", "Location", "Link", "Notes"]
    df_links  = pd.DataFrame(search_links, columns=link_cols)

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df_jobs.to_excel(writer,  index=False, sheet_name="All Jobs")
        df_links.to_excel(writer, index=False, sheet_name="Job Board Links")

        # ── Style: All Jobs ────────────────────────────────────────────────
        ws_jobs = writer.sheets["All Jobs"]
        tier_idx = COLUMNS.index("Location Tier") + 1

        def jobs_fill(row):
            tier = row[tier_idx - 1].value
            return green if tier == "A" else yellow

        _style_sheet(ws_jobs, link_col_letter="H", fill_fn=jobs_fill)

        job_col_widths = {
            "Source": 22, "Job Title": 52, "Company": 32, "Salary": 18,
            "Location": 26, "Work Type": 10, "Location Tier": 13,
            "Application Link": 58, "Documents Required": 38, "Date Scraped": 13,
        }
        for col in ws_jobs.columns:
            ws_jobs.column_dimensions[col[0].column_letter].width = job_col_widths.get(col[0].value, 18)
        ws_jobs.freeze_panes = "A2"

        # ── Style: Job Board Links ─────────────────────────────────────────
        ws_links = writer.sheets["Job Board Links"]
        _style_sheet(ws_links, link_col_letter="D", fill_fn=lambda row: blue)

        link_col_widths = {
            "Job Board": 35, "Search Query": 32, "Location": 22,
            "Link": 70, "Notes": 55,
        }
        for col in ws_links.columns:
            ws_links.column_dimensions[col[0].column_letter].width = link_col_widths.get(col[0].value, 20)
        ws_links.freeze_panes = "A2"

    log.info("Saved %d job rows + %d board links to %s", len(df_jobs), len(df_links), out_path)
    return out_path


# ── Main ───────────────────────────────────────────────────────────────────

def run_scraper():
    log.info("=" * 70)
    log.info("Job scraper started — %s", datetime.now().strftime("%Y-%m-%d %H:%M"))

    missing_keys = []
    if not ADZUNA_APP_ID:  missing_keys.append("ADZUNA_APP_ID / ADZUNA_APP_KEY (https://developer.adzuna.com/)")
    if not USAJOBS_KEY:    missing_keys.append("USAJOBS_API_KEY / USAJOBS_EMAIL (https://developer.usajobs.gov/)")
    if not JOOBLE_KEY:     missing_keys.append("JOOBLE_API_KEY (https://jooble.org/api/about)")
    if not SERPAPI_KEY:    missing_keys.append("SERPAPI_KEY (https://serpapi.com/)")
    if missing_keys:
        log.warning(
            "The following free API keys are not configured — those sources will be skipped:\n%s\n"
            "Set them as environment variables or paste them at the top of job_scraper.py",
            "\n".join(f"  • {k}" for k in missing_keys)
        )

    all_jobs: list[dict] = []

    for query in QUERIES:
        log.info("── %s ──", query)
        for fn in (scrape_adzuna, scrape_usajobs, scrape_jooble, scrape_jobsacuk, scrape_serpapi):
            try:
                all_jobs.extend(fn(query))
            except Exception as exc:
                log.error("%s / [%s] crashed: %s", fn.__name__, query, exc)
        time.sleep(1)

    all_jobs      = dedupe(all_jobs)
    search_links  = build_search_links()
    log.info("Real job listings: %d  |  Board search links: %d", len(all_jobs), len(search_links))

    out = save_excel(all_jobs, search_links)
    n   = len(all_jobs)
    print(f"\n{'=' * 60}")
    print(f"  Real listings found : {n}")
    print(f"  Board search links  : {len(search_links)}")
    print(f"  Spreadsheet saved to: {out}")
    print(f"{'=' * 60}\n")
    if n == 0:
        print("  WARNING: No real jobs scraped yet -- add API keys to run_scraper.bat")
        print("     then re-run.  Use the 'Job Board Links' sheet to search manually.\n")
    log.info("Finished — %s", datetime.now().strftime("%Y-%m-%d %H:%M"))


if __name__ == "__main__":
    run_scraper()
